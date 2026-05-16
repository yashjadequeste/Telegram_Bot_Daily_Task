import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

TEMPLATE_PATH = Path("templates/Daily Task.xlsx")
WORKBOOK_PATH = Path("reports/Daily_Report.xlsx")
DATE_FORMAT = "d/m/yyyy"

FMT_DATE = 2
FMT_TASK = 3
FMT_SPACER = 4

NO_FILL = PatternFill(fill_type=None)

_template_sheet_cache = None


def get_template_sheet():
    global _template_sheet_cache
    if _template_sheet_cache is None:
        twb = load_workbook(TEMPLATE_PATH)
        name = get_month_sheet_name()
        _template_sheet_cache = (
            twb[name] if name in twb.sheetnames else twb[twb.sheetnames[0]]
        )
    return _template_sheet_cache


def get_month_sheet_name(dt=None):
    dt = dt or datetime.now()
    return dt.strftime("%B %Y")


def get_workbook():
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not WORKBOOK_PATH.exists():
        shutil.copy(TEMPLATE_PATH, WORKBOOK_PATH)
    return load_workbook(WORKBOOK_PATH)


def save_workbook(wb):
    wb.save(WORKBOOK_PATH)
    return str(WORKBOOK_PATH)


def clear_sheet_data(sheet):
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)


def ensure_month_sheet(wb):
    name = get_month_sheet_name()
    if name in wb.sheetnames:
        wb.active = wb[name]
        return wb[name]

    source = wb[wb.sheetnames[0]]
    new_sheet = wb.copy_worksheet(source)
    new_sheet.title = name[:31]
    clear_sheet_data(new_sheet)
    wb.active = new_sheet
    return new_sheet


def copy_cell_format(source_sheet, source_row, target_sheet, target_row, copy_fill=True):
    for col in range(1, target_sheet.max_column + 1):
        source = source_sheet.cell(source_row, col)
        target = target_sheet.cell(target_row, col)
        target.font = copy(source.font)
        if copy_fill:
            target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format

    if source_row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[target_row].height = (
            source_sheet.row_dimensions[source_row].height
        )


def row_is_empty(sheet, row):
    if row < 1:
        return True
    for col in range(1, 5):
        value = sheet.cell(row, col).value
        if value is not None and str(value).strip() != "":
            return False
    return True


def unmerge_row(sheet, row):
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row:
            try:
                sheet.unmerge_cells(str(merged))
            except KeyError:
                pass


def clear_stale_merges(sheet):
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row > 1:
            try:
                sheet.unmerge_cells(str(merged))
            except KeyError:
                pass


def clear_row_fill(sheet, row):
    for col in range(1, 5):
        sheet.cell(row, col).fill = NO_FILL


def write_spacer_row(sheet, row):
    tsh = get_template_sheet()
    unmerge_row(sheet, row)
    copy_cell_format(tsh, FMT_SPACER, sheet, row, copy_fill=False)
    for col in range(1, 5):
        sheet.cell(row, col).value = None
    clear_row_fill(sheet, row)


def write_date_row(sheet, row, date_value):
    tsh = get_template_sheet()
    unmerge_row(sheet, row)
    copy_cell_format(tsh, FMT_DATE, sheet, row)
    cell = sheet.cell(row, 1)
    cell.value = date_value
    cell.number_format = DATE_FORMAT
    for col in range(2, 5):
        sheet.cell(row, col).value = None
    merge_date_row(sheet, row)


def write_task_row(sheet, row, task_text, task_type, status):
    tsh = get_template_sheet()
    unmerge_row(sheet, row)
    copy_cell_format(tsh, FMT_TASK, sheet, row)
    sheet.cell(row, 1).value = 1
    sheet.cell(row, 2).value = task_text
    sheet.cell(row, 3).value = task_type
    sheet.cell(row, 4).value = status
    for col in range(2, 5):
        sheet.cell(row, col).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def merge_date_row(sheet, row):
    range_name = f"A{row}:D{row}"
    for merged in list(sheet.merged_cells.ranges):
        if str(merged) == range_name:
            return
    sheet.merge_cells(range_name)


def collect_blocks(sheet):
    blocks = []
    row = 2
    while row <= sheet.max_row:
        value = sheet.cell(row, 1).value
        if isinstance(value, datetime):
            task_row = row + 1
            task_text = ""
            task_type = ""
            status = ""
            if task_row <= sheet.max_row and sheet.cell(task_row, 1).value == 1:
                task_text = sheet.cell(task_row, 2).value or ""
                task_type = sheet.cell(task_row, 3).value or ""
                status = sheet.cell(task_row, 4).value or ""
            blocks.append((value, task_text, task_type, status))
            row += 1
            while row <= sheet.max_row and not isinstance(
                sheet.cell(row, 1).value, datetime
            ):
                row += 1
        else:
            row += 1
    return blocks


def normalize_sheet(sheet):
    """Each day: date (grey) -> task -> ONE white spacer. No double gaps."""
    blocks = collect_blocks(sheet)

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    clear_stale_merges(sheet)

    task_row_result = None
    for date_val, task_text, task_type, status in blocks:
        write_date_row(sheet, sheet.max_row + 1, date_val)
        task_row = sheet.max_row + 1
        write_task_row(sheet, task_row, task_text, task_type, status)
        write_spacer_row(sheet, sheet.max_row + 1)

        if date_val.date() == datetime.now().date():
            task_row_result = task_row

    return task_row_result


def find_today_block(sheet):
    today = datetime.now().date()
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if isinstance(value, datetime) and value.date() == today:
            return row, row + 1, row + 2
    return None


def append_day_block(sheet, task_text, task_type, status):
    if sheet.max_row >= 2 and not row_is_empty(sheet, sheet.max_row):
        write_spacer_row(sheet, sheet.max_row + 1)

    write_date_row(sheet, sheet.max_row + 1, datetime.now().replace(
        hour=0, minute=0, second=0, microsecond=0
    ))
    task_row = sheet.max_row + 1
    write_task_row(sheet, task_row, task_text, task_type, status)
    write_spacer_row(sheet, sheet.max_row + 1)
    return task_row


def update_excel(task_text, task_type, status):
    wb = get_workbook()
    sheet = ensure_month_sheet(wb)

    block = find_today_block(sheet)
    if block:
        _, task_row, _ = block
        write_task_row(sheet, task_row, task_text, task_type, status)
        normalize_sheet(sheet)
        block = find_today_block(sheet)
        task_row = block[1] if block else task_row
    else:
        append_day_block(sheet, task_text, task_type, status)
        block = find_today_block(sheet)
        task_row = block[1] if block else sheet.max_row

    return save_workbook(wb), task_row


def update_existing_status(row, status):
    wb = get_workbook()
    sheet = ensure_month_sheet(wb)

    task_text = sheet.cell(row, 2).value or ""
    task_type = sheet.cell(row, 3).value or ""
    write_task_row(sheet, row, task_text, task_type, status)
    normalize_sheet(sheet)
    return save_workbook(wb)


def complete_rows_in_excel(rows):
    wb = get_workbook()
    sheet = ensure_month_sheet(wb)

    for row in rows:
        if row and row >= 2:
            task_text = sheet.cell(row, 2).value or ""
            task_type = sheet.cell(row, 3).value or ""
            write_task_row(sheet, row, task_text, task_type, "Completed")

    normalize_sheet(sheet)
    return save_workbook(wb)
